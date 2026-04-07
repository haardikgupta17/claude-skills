#!/usr/bin/env python3
"""
Sprint Task Tracker — stdin JSON interface for sprint-pilot.

Usage:
    echo '<json>' | python3 update_sprint_tasks.py --start --sprint "SPRINT NAME"
    echo '<json>' | python3 update_sprint_tasks.py --end   --sprint "SPRINT NAME"
    echo '<json>' | python3 update_sprint_tasks.py --summon --sprint "SPRINT NAME"
    python3 update_sprint_tasks.py --dashboard

JSON format (array of ticket objects):
  [
    {
      "key": "PROJ-XXXXX",
      "summary": "...",
      "priority": "P2-High",
      "status": "Dev Assigned",
      "action_brief": "...",       (optional — AI-generated next steps)
      "description": "...",        (optional)
      "latest_comment": "..."      (optional)
    },
    ...
  ]
"""

import sys
import json
import argparse
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = Path(__file__).parent / "Sprint_Tasks.xlsx"
VELOCITY_SHEET = "Velocity"

PRIORITY_ORDER = {
    "P1-Critical":     1,
    "P2-High":         2,
    "P3-Medium":       3,
    "P4-Low":          4,
    "P5-Undetermined": 5,
}

PRIORITY_COLORS = {
    "P1-Critical":     "FF0000",
    "P2-High":         "FF6600",
    "P3-Medium":       "FFC000",
    "P4-Low":          "92D050",
    "P5-Undetermined": "808080",
}

STATUS_COLORS = {
    "Blocked":              "FF0000",
    "Dev Assigned":         "4472C4",
    "Internal Review":      "70AD47",
    "Ready for Deployment": "00B0F0",
    "New":                  "808080",
    "In Progress":          "4472C4",
    "Resolved":             "00B050",
    "Closed":               "00B050",
    "Done":                 "00B050",
}

DONE_STATUSES = {"Resolved", "Closed", "Done"}

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 9-column layout
HEADERS = [
    "S.No", "Ticket Number", "Summary", "Priority", "Status",
    "Action Brief", "Carry-Over", "Risk / Dependency Flag", "Notes",
]
COL_WIDTHS = [6, 16, 55, 18, 22, 65, 12, 30, 30]

# Column constants (1-based)
COL_SNO      = 1
COL_KEY      = 2
COL_SUMMARY  = 3
COL_PRIORITY = 4
COL_STATUS   = 5
COL_ACTION   = 6
COL_CARRY    = 7
COL_RISK     = 8
COL_NOTES    = 9

# Velocity sheet columns
V_HEADERS = [
    "Sprint", "Start Date", "End Date",
    "Planned", "Carry-Overs In", "Scope Additions",
    "Resolved", "Blocked", "Remaining",
    "Resolution Rate %", "Carry-Over Rate %", "Scope Creep Rate %",
]
V_COL_WIDTHS = [32, 14, 14, 10, 14, 16, 10, 10, 11, 18, 18, 18]
V_COL = {name: idx for idx, name in enumerate(V_HEADERS, 1)}


def priority_key(ticket):
    return PRIORITY_ORDER.get(ticket.get("priority", ""), 99)


def get_last_sprint_sheet_name_and_keys(wb):
    """Return (last_sprint_sheet_name, set_of_ticket_keys), skipping non-sprint sheets."""
    skip = {VELOCITY_SHEET, "Dashboard"}
    sprint_sheets = [n for n in wb.sheetnames if n not in skip]
    if not sprint_sheets:
        return None, set()
    last_name = sprint_sheets[-1]
    ws = wb[last_name]
    keys = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        val = row[1]
        if val:
            keys.add(str(val).strip())
    return last_name, keys


# -- Velocity sheet helpers ----------------------------------------------------

def _ensure_velocity_sheet(wb):
    if VELOCITY_SHEET in wb.sheetnames:
        return wb[VELOCITY_SHEET]

    ws = wb.create_sheet(title=VELOCITY_SHEET, index=0)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col_idx, (header, width) in enumerate(zip(V_HEADERS, V_COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return ws


def _find_velocity_row(ws, sprint_name):
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=V_COL["Sprint"]).value
        if val and str(val).strip() == sprint_name.strip():
            return row_idx
    return None


def _rate_cell(ws, row, col, numerator, denominator):
    pct = round(numerator / denominator * 100, 1) if denominator and denominator > 0 else None
    c = ws.cell(row=row, column=col, value=f"{pct}%" if pct is not None else "\u2014")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
    if pct is not None:
        if col == V_COL["Resolution Rate %"]:
            color = "00B050" if pct >= 70 else ("FFC000" if pct >= 40 else "FF0000")
        else:
            color = "00B050" if pct <= 20 else ("FFC000" if pct <= 40 else "FF0000")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return c


def _write_velocity_row(ws, row, sprint_name, start_dt, end_dt,
                         planned, carry_overs_in, scope_adds,
                         resolved, blocked, remaining):
    alt = row % 2 == 0
    bg = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid") if alt else None

    def plain(col, val, bold=False, center=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold)
        c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
        c.border = BORDER
        if bg:
            c.fill = bg

    plain(V_COL["Sprint"], sprint_name, bold=True, center=False)
    plain(V_COL["Start Date"], start_dt)
    plain(V_COL["End Date"], end_dt if end_dt else "\u2014")

    for col, val in [
        (V_COL["Planned"],         planned),
        (V_COL["Carry-Overs In"],  carry_overs_in),
        (V_COL["Scope Additions"], scope_adds if scope_adds is not None else "\u2014"),
        (V_COL["Resolved"],        resolved if resolved is not None else "\u2014"),
        (V_COL["Blocked"],         blocked if blocked is not None else "\u2014"),
        (V_COL["Remaining"],       remaining if remaining is not None else "\u2014"),
    ]:
        plain(col, val)

    if resolved is not None:
        total = (resolved or 0) + (blocked or 0) + (remaining or 0)
        _rate_cell(ws, row, V_COL["Resolution Rate %"], resolved or 0, total)
        _rate_cell(ws, row, V_COL["Carry-Over Rate %"], carry_overs_in or 0, planned or 1)
        _rate_cell(ws, row, V_COL["Scope Creep Rate %"], scope_adds or 0, planned or 1)
    else:
        for col in (V_COL["Resolution Rate %"], V_COL["Carry-Over Rate %"], V_COL["Scope Creep Rate %"]):
            c = ws.cell(row=row, column=col, value="In Progress")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = BORDER
            c.font = Font(italic=True, color="808080")
            if bg:
                c.fill = bg

    ws.row_dimensions[row].height = 18


def velocity_record_start(wb, sprint_name, planned, carry_overs_in):
    vs = _ensure_velocity_sheet(wb)
    row = _find_velocity_row(vs, sprint_name)
    if row is None:
        row = vs.max_row + 1 if vs.max_row >= 2 else 2
        if vs.max_row == 1:
            row = 2
    _write_velocity_row(vs, row, sprint_name, start_dt=date.today().isoformat(), end_dt=None,
                         planned=planned, carry_overs_in=carry_overs_in,
                         scope_adds=None, resolved=None, blocked=None, remaining=None)
    print(f"Velocity: recorded sprint start \u2014 {planned} planned, {carry_overs_in} carry-overs in.")


def velocity_record_end(wb, sprint_name, scope_adds, resolved, blocked, remaining):
    vs = _ensure_velocity_sheet(wb)
    row = _find_velocity_row(vs, sprint_name)
    if row is not None:
        planned = vs.cell(row=row, column=V_COL["Planned"]).value or 0
        carry_overs_in = vs.cell(row=row, column=V_COL["Carry-Overs In"]).value or 0
        start_dt = vs.cell(row=row, column=V_COL["Start Date"]).value or "\u2014"
    else:
        row = vs.max_row + 1 if vs.max_row >= 2 else 2
        planned = resolved + blocked + remaining + scope_adds
        carry_overs_in = 0
        start_dt = "\u2014"

    _write_velocity_row(vs, row, sprint_name, start_dt=start_dt, end_dt=date.today().isoformat(),
                         planned=planned, carry_overs_in=carry_overs_in, scope_adds=scope_adds,
                         resolved=resolved, blocked=blocked, remaining=remaining)

    resolution_rate = round(resolved / max(resolved + blocked + remaining, 1) * 100, 1)
    print(f"Velocity: sprint closed \u2014 {resolved} resolved ({resolution_rate}% rate), "
          f"{scope_adds} scope adds, {blocked} blocked.")


# -- Row height auto-fit -------------------------------------------------------

def _auto_row_height(ws, row_idx, min_height=20, line_height=15):
    """Estimate row height based on longest wrapped text in the row."""
    max_lines = 1
    for col_idx in range(1, len(HEADERS) + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if not val:
            continue
        text = str(val)
        col_width = COL_WIDTHS[col_idx - 1]
        chars_per_line = max(int(col_width * 1.2), 1)
        lines = max(len(text) // chars_per_line + 1, text.count('\n') + 1)
        max_lines = max(max_lines, lines)
    ws.row_dimensions[row_idx].height = max(min_height, max_lines * line_height)


# -- Sprint sheet: --start mode ------------------------------------------------

def add_sprint_sheet_start(wb, sprint_name, tickets):
    sheet_name = sprint_name[:31]

    last_sheet_name, last_keys = get_last_sprint_sheet_name_and_keys(wb)

    for t in tickets:
        t["carried_over"] = t["key"] in last_keys

    carried = sorted([t for t in tickets if t["carried_over"]], key=priority_key)
    fresh = sorted([t for t in tickets if not t["carried_over"]], key=priority_key)
    sorted_tickets = carried + fresh

    if sheet_name in wb.sheetnames:
        print(f"Sheet '{sheet_name}' already exists \u2014 overwriting.")
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Data rows
    action_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for i, t in enumerate(sorted_tickets, 1):
        row = i + 1

        is_carried = t.get("carried_over", False)
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if i % 2 == 0 else None

        # S.No
        c = ws.cell(row=row, column=COL_SNO, value=i)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Ticket key
        c = ws.cell(row=row, column=COL_KEY, value=t["key"])
        c.font = Font(bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Summary
        c = ws.cell(row=row, column=COL_SUMMARY, value=t["summary"])
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Priority
        p_color = PRIORITY_COLORS.get(t.get("priority", ""), "808080")
        c = ws.cell(row=row, column=COL_PRIORITY, value=t.get("priority", ""))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=p_color, end_color=p_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Status
        s_color = STATUS_COLORS.get(t.get("status", ""), "808080")
        c = ws.cell(row=row, column=COL_STATUS, value=t.get("status", ""))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=s_color, end_color=s_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Action Brief (AI-generated)
        action_text = t.get("action_brief", "")
        c = ws.cell(row=row, column=COL_ACTION, value=action_text)
        c.font = Font(color="1A3353", size=10)
        c.fill = action_fill
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER

        # Carry-Over flag (Y/N)
        c = ws.cell(row=row, column=COL_CARRY, value="Y" if is_carried else "N")
        c.font = Font(bold=True, color="C00000" if is_carried else "808080")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Risk / Dependency Flag (user's own — always empty)
        c = ws.cell(row=row, column=COL_RISK, value="")
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Notes (user's own — always empty)
        c = ws.cell(row=row, column=COL_NOTES, value="")
        c.border = BORDER
        if alt_fill:
            c.fill = alt_fill

        # Auto-fit row height based on content
        _auto_row_height(ws, row)

    ws.freeze_panes = "A2"

    carried_count = len(carried)
    print(f"Added sheet: '{sheet_name}' | Total: {len(sorted_tickets)} | "
          f"Carried over: {carried_count} | Fresh: {len(fresh)}")

    velocity_record_start(wb, sprint_name, planned=len(sorted_tickets), carry_overs_in=carried_count)
    return carried_count


# -- Sprint sheet: --end mode --------------------------------------------------

def update_sprint_sheet_end(wb, sprint_name, tickets):
    sheet_name = sprint_name[:31]

    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for name in wb.sheetnames:
            if name not in {VELOCITY_SHEET, "Dashboard"} and name.strip().lower() == sheet_name.strip().lower():
                ws = wb[name]
                sheet_name = name
                break

    if ws is None:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    # Build key -> row map from column B
    key_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=COL_KEY).value
        if val:
            key_to_row[str(val).strip()] = row_idx

    resolved = blocked = carried = 0
    scope_creep_tickets = []

    ticket_map = {t["key"]: t for t in tickets}

    # Update existing rows — only touch Status (col 5), preserve Action Brief & Notes
    for key, t in ticket_map.items():
        row_idx = key_to_row.get(key)
        if row_idx is None:
            scope_creep_tickets.append(t)
            continue

        status = t.get("status", "")
        s_color = STATUS_COLORS.get(status, "808080")
        c = ws.cell(row=row_idx, column=COL_STATUS, value=status)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=s_color, end_color=s_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        if status in DONE_STATUSES:
            resolved += 1
        elif status == "Blocked":
            blocked += 1
        else:
            carried += 1

    # Append scope-creep tickets at the bottom
    scope_creep_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    action_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    scope_sorted = sorted(scope_creep_tickets, key=priority_key)
    next_row = ws.max_row + 1
    existing_count = next_row - 2

    for i, t in enumerate(scope_sorted, 1):
        row = next_row + i - 1
        serial = existing_count + i
        ws.row_dimensions[row].height = 28

        # S.No
        c = ws.cell(row=row, column=COL_SNO, value=serial)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Ticket key
        c = ws.cell(row=row, column=COL_KEY, value=t["key"])
        c.font = Font(bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Summary
        c = ws.cell(row=row, column=COL_SUMMARY, value=t.get("summary", ""))
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.fill = scope_creep_fill
        c.border = BORDER

        # Priority
        p_color = PRIORITY_COLORS.get(t.get("priority", ""), "808080")
        c = ws.cell(row=row, column=COL_PRIORITY, value=t.get("priority", ""))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=p_color, end_color=p_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Status
        status = t.get("status", "")
        s_color = STATUS_COLORS.get(status, "808080")
        c = ws.cell(row=row, column=COL_STATUS, value=status)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=s_color, end_color=s_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Action Brief (for scope additions, write if provided)
        action_text = t.get("action_brief", "")
        c = ws.cell(row=row, column=COL_ACTION, value=action_text)
        c.font = Font(color="1A3353", size=10)
        c.fill = action_fill
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER

        # Carry-Over flag (scope additions are not carry-overs)
        c = ws.cell(row=row, column=COL_CARRY, value="N")
        c.font = Font(bold=True, color="808080")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Risk / Dependency Flag — mark scope addition here
        c = ws.cell(row=row, column=COL_RISK, value="\U0001f4cc Scope Addition")
        c.font = Font(bold=True, color="7B5A00")
        c.fill = scope_creep_fill
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER

        # Notes
        c = ws.cell(row=row, column=COL_NOTES, value="")
        c.fill = scope_creep_fill
        c.border = BORDER

        # Auto-fit row height
        _auto_row_height(ws, row)

        if status in DONE_STATUSES:
            resolved += 1
        elif status == "Blocked":
            blocked += 1
        else:
            carried += 1

    scope_count = len(scope_creep_tickets)
    print(f"Sprint closed. Resolved: {resolved} \u2705 | Blocked: {blocked} \u274c | "
          f"Remaining: {carried} \u26a0 | Scope additions: {scope_count} \U0001f4cc")

    velocity_record_end(wb, sprint_name, scope_adds=scope_count,
                        resolved=resolved, blocked=blocked, remaining=carried)
    return resolved, blocked, carried, scope_count


# -- Sprint sheet: --summon mode ------------------------------------------------

def summon_ticket(wb, sprint_name, tickets):
    """Append one or more tickets to an existing sprint sheet as scope additions."""
    sheet_name = sprint_name[:31]

    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for name in wb.sheetnames:
            if name not in {VELOCITY_SHEET, "Dashboard"} and name.strip().lower() == sheet_name.strip().lower():
                ws = wb[name]
                sheet_name = name
                break

    if ws is None:
        print(f"ERROR: Sheet '{sheet_name}' not found. Run --start (liftoff) first.", file=sys.stderr)
        sys.exit(1)

    # Check for duplicates
    existing_keys = set()
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=COL_KEY).value
        if val:
            existing_keys.add(str(val).strip())

    scope_creep_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    action_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    next_row = ws.max_row + 1
    existing_count = next_row - 2
    added = 0

    for t in tickets:
        if t["key"] in existing_keys:
            print(f"Skipped {t['key']} — already in sheet.")
            continue

        added += 1
        row = next_row + added - 1
        serial = existing_count + added

        # S.No
        c = ws.cell(row=row, column=COL_SNO, value=serial)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Ticket key
        c = ws.cell(row=row, column=COL_KEY, value=t["key"])
        c.font = Font(bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Summary
        c = ws.cell(row=row, column=COL_SUMMARY, value=t.get("summary", ""))
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.fill = scope_creep_fill
        c.border = BORDER

        # Priority
        p_color = PRIORITY_COLORS.get(t.get("priority", ""), "808080")
        c = ws.cell(row=row, column=COL_PRIORITY, value=t.get("priority", ""))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=p_color, end_color=p_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Status
        status = t.get("status", "")
        s_color = STATUS_COLORS.get(status, "808080")
        c = ws.cell(row=row, column=COL_STATUS, value=status)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=s_color, end_color=s_color, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER

        # Action Brief
        action_text = t.get("action_brief", "")
        c = ws.cell(row=row, column=COL_ACTION, value=action_text)
        c.font = Font(color="1A3353", size=10)
        c.fill = action_fill
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER

        # Carry-Over (scope additions are not carry-overs)
        c = ws.cell(row=row, column=COL_CARRY, value="N")
        c.font = Font(bold=True, color="808080")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = scope_creep_fill
        c.border = BORDER

        # Risk / Dependency Flag
        c = ws.cell(row=row, column=COL_RISK, value="\U0001f4cc Scope Addition")
        c.font = Font(bold=True, color="7B5A00")
        c.fill = scope_creep_fill
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BORDER

        # Notes
        c = ws.cell(row=row, column=COL_NOTES, value="")
        c.fill = scope_creep_fill
        c.border = BORDER

        # Auto-fit row height
        _auto_row_height(ws, row)

    print(f"Summoned {added} ticket(s) into '{sheet_name}' \U0001f4cc")
    return added


# -- Dashboard generation ------------------------------------------------------

DASHBOARD_PNG = EXCEL_PATH.parent / "Sprint_Velocity_Dashboard.png"

def generate_dashboard(wb):
    """Generate 4-panel velocity chart and embed in Dashboard sheet."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import numpy as np
        from openpyxl.drawing.image import Image as XlImage
    except ImportError:
        print("WARNING: matplotlib not installed — skipping dashboard generation.")
        print("  Install with: pip install matplotlib")
        return

    vs = wb[VELOCITY_SHEET] if VELOCITY_SHEET in wb.sheetnames else None
    if vs is None or vs.max_row < 2:
        print("WARNING: No velocity data — skipping dashboard.")
        return

    sprints, planned, carry_overs, scope_adds = [], [], [], []
    resolved, blocked, remaining_ = [], [], []

    for row in range(2, vs.max_row + 1):
        name = vs.cell(row, V_COL["Sprint"]).value
        if not name:
            continue
        label = str(name)
        if len(label) > 20:
            label = label[:18] + ".."
        sprints.append(label)
        planned.append(vs.cell(row, V_COL["Planned"]).value or 0)
        carry_overs.append(vs.cell(row, V_COL["Carry-Overs In"]).value or 0)
        sa = vs.cell(row, V_COL["Scope Additions"]).value
        scope_adds.append(sa if isinstance(sa, (int, float)) else 0)
        r = vs.cell(row, V_COL["Resolved"]).value
        resolved.append(r if isinstance(r, (int, float)) else 0)
        b = vs.cell(row, V_COL["Blocked"]).value
        blocked.append(b if isinstance(b, (int, float)) else 0)
        rm = vs.cell(row, V_COL["Remaining"]).value
        remaining_.append(rm if isinstance(rm, (int, float)) else 0)

    if not sprints:
        print("WARNING: No sprint data in Velocity sheet — skipping dashboard.")
        return

    # Compute rates
    res_rate, co_rate, sc_rate = [], [], []
    for i in range(len(sprints)):
        total = (resolved[i] + blocked[i] + remaining_[i]) or 1
        res_rate.append(round(resolved[i] / total * 100, 1))
        co_rate.append(round(carry_overs[i] / max(planned[i], 1) * 100, 1))
        sc_rate.append(round(scope_adds[i] / max(planned[i], 1) * 100, 1))

    plt.style.use("dark_background")
    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle("Sprint Velocity Dashboard", fontsize=18, fontweight="bold", color="#00BFFF", y=0.97)
    fig.patch.set_facecolor("#1a1a2e")
    for ax in axes.flat:
        ax.set_facecolor("#16213e")

    x = np.arange(len(sprints))
    w = 0.25

    # Panel 1: Sprint Outcome Breakdown
    ax1 = axes[0, 0]
    ax1.bar(x - w, resolved, w, label="Resolved", color="#00B050")
    ax1.bar(x, blocked, w, label="Blocked", color="#FF4444")
    ax1.bar(x + w, remaining_, w, label="Remaining", color="#FFC000")
    ax1.set_title("Sprint Outcome Breakdown", fontsize=12, fontweight="bold")
    ax1.set_xticks(x)
    ax1.set_xticklabels(sprints, fontsize=8, rotation=15)
    ax1.legend(fontsize=9)
    ax1.set_ylabel("Tickets")

    # Panel 2: Resolution Rate Trend
    ax2 = axes[0, 1]
    ax2.plot(sprints, res_rate, "o-", color="#00BFFF", linewidth=2, markersize=8)
    ax2.axhline(y=70, color="#00B050", linestyle="--", alpha=0.5, label="Target 70%")
    ax2.fill_between(sprints, res_rate, alpha=0.15, color="#00BFFF")
    ax2.set_title("Resolution Rate Trend", fontsize=12, fontweight="bold")
    ax2.set_ylabel("Resolution %")
    ax2.set_ylim(0, 105)
    ax2.legend(fontsize=9)
    for i, v in enumerate(res_rate):
        ax2.annotate(f"{v}%", (sprints[i], v), textcoords="offset points", xytext=(0, 10), ha="center", fontsize=9)

    # Panel 3: Planned vs Actual Scope
    ax3 = axes[1, 0]
    actual = [p + s for p, s in zip(planned, scope_adds)]
    ax3.bar(x - 0.15, planned, 0.3, label="Planned", color="#4472C4")
    ax3.bar(x + 0.15, actual, 0.3, label="Actual (+ scope adds)", color="#FF6600")
    ax3.set_title("Planned vs Actual Scope", fontsize=12, fontweight="bold")
    ax3.set_xticks(x)
    ax3.set_xticklabels(sprints, fontsize=8, rotation=15)
    ax3.legend(fontsize=9)
    ax3.set_ylabel("Tickets")

    # Panel 4: Scope Creep Rate
    ax4 = axes[1, 1]
    ax4.plot(sprints, sc_rate, "s-", color="#FF6600", linewidth=2, markersize=8)
    ax4.axhline(y=20, color="#FF4444", linestyle="--", alpha=0.5, label="Danger 20%")
    ax4.fill_between(sprints, sc_rate, alpha=0.15, color="#FF6600")
    ax4.set_title("Scope Creep Rate", fontsize=12, fontweight="bold")
    ax4.set_ylabel("Scope Creep %")
    ax4.set_ylim(0, max(max(sc_rate) + 15, 35) if sc_rate else 35)
    ax4.legend(fontsize=9)
    for i, v in enumerate(sc_rate):
        ax4.annotate(f"{v}%", (sprints[i], v), textcoords="offset points", xytext=(0, 10), ha="center", fontsize=9)

    plt.tight_layout(rect=[0, 0, 1, 0.94])

    png_path = EXCEL_PATH.parent / (EXCEL_PATH.stem + "_Dashboard.png")
    fig.savefig(png_path, dpi=150, bbox_inches="tight")
    plt.close(fig)

    # Embed in Dashboard sheet
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ds = wb.create_sheet("Dashboard", 0)
    ds.sheet_properties.tabColor = "1F4E79"
    ds["A1"] = "Sprint Velocity Dashboard"
    ds["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ds.row_dimensions[1].height = 30
    from openpyxl.drawing.image import Image as XlImage
    img = XlImage(str(png_path))
    img.width = 1100
    img.height = 690
    ds.add_image(img, "A3")

    print(f"Dashboard: chart saved to {png_path} and embedded in Excel.")


# -- Entry point ---------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Sprint task tracker for sprint-pilot")
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument("--start", action="store_true", help="Sprint start mode")
    mode_group.add_argument("--end", action="store_true", help="Sprint end mode")
    mode_group.add_argument("--summon", action="store_true", help="Add ticket(s) mid-sprint as scope additions")
    mode_group.add_argument("--dashboard", action="store_true", help="Regenerate velocity dashboard only")
    parser.add_argument("--sprint", required=False, help="Sprint name (sheet name, max 31 chars)")
    args = parser.parse_args()

    if args.dashboard:
        if not EXCEL_PATH.exists():
            print("ERROR: Excel file not found. Run --start first.", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        generate_dashboard(wb)
        wb.save(EXCEL_PATH)
        print(f"Saved: {EXCEL_PATH}")
        return

    if not args.sprint:
        print("ERROR: --sprint is required for --start, --end, and --summon modes.", file=sys.stderr)
        sys.exit(1)

    raw = sys.stdin.read().strip()
    if not raw:
        print("ERROR: No JSON received on stdin.", file=sys.stderr)
        sys.exit(1)

    try:
        tickets = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON on stdin: {e}", file=sys.stderr)
        sys.exit(1)

    if not isinstance(tickets, list):
        print("ERROR: JSON must be a list of ticket objects.", file=sys.stderr)
        sys.exit(1)

    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        print(f"Opened: {EXCEL_PATH.name}")
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print(f"Creating new file: {EXCEL_PATH.name}")

    if args.start:
        add_sprint_sheet_start(wb, args.sprint, tickets)
    elif args.end:
        update_sprint_sheet_end(wb, args.sprint, tickets)
        generate_dashboard(wb)
    elif args.summon:
        summon_ticket(wb, args.sprint, tickets)

    wb.save(EXCEL_PATH)
    print(f"Saved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
